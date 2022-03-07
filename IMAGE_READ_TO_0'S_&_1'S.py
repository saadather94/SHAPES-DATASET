import cv2
import openpyxl
import glob

##################################################################################

path = "C:\\Users\\RAPTOR\\Desktop\\circle dataset\\*"

for file in glob.glob(path):

    print(file)
    img = cv2.imread(file,2)

    ret, bw_img = cv2.threshold(img,127,255,cv2.THRESH_BINARY)


    pix_val = list(bw_img)
    pix_val_flat = [x for sets in pix_val for x in sets]
    #print(pix_val_flat,'\n')

    list_ = []
    for val in pix_val_flat:
        if (val<=255 and val>=200):
            list_.append(0)
        elif(val<30):
            list_.append(1)

    print(list_,"\n")

    file = 'C:\\Users\\RAPTOR\\Desktop\\one.xlsx'

    wb = openpyxl.load_workbook(filename=file)
    ws = wb['Sheet1']     # .get_sheet_by_name('Sheet1')
    row = ws.max_row +1


    for col, entry in enumerate(list_, start=1):
        ws.cell(row=row, column=col, value=entry)

    wb.save(file)